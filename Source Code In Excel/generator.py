import hashlib, os, time

from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment

def get_relative_path(subdirectory, directory):
    directory = subdirectory.replace(directory, "")
    if len(directory) > 0:
        directory = '.' + directory
    else:
        directory = ".\\"
    return directory

def get_path():
    path = input("Enter path to directory: ")
    if os.path.exists(path):
        return path
    return get_path()

# Returns [(flledirectory, filename)]
def get_all_files_data(path):
    filenames = []
    for x in os.walk(path):
        new_filenames = []
        for y in x[2]:
            if y.endswith(".cs"):
                new_filenames.append(y)
        if len(new_filenames) != 0:
            filetuple = (x[0], new_filenames)
            filenames.append(filetuple)
    return filenames

# Returns [(flledirectory, [(counter, name, creationtime, bytesize, linescount, checksum)] )]
def create_xlsx_data(sourcepath):
    filetable = []
    counter = 1
    for filetuple in filenames:
        relative_directory = get_relative_path(filetuple[0], sourcepath)
        filebind = (relative_directory, [])
        
        for name in filetuple[1]:
            filepath = os.path.join(filetuple[0], name)
            file = open(filepath, "r", errors="ignore")
            filestats = os.stat(filepath)
            filedata = file.read()
            
            creationtime = os.path.getctime(filepath)
            creationtime = time.ctime(creationtime)
            creationtime = time.strptime(creationtime)
            creationtime = time.strftime("%d.%m.%Y %H:%M", creationtime)

            bytesize = filestats.st_size
            linescount = filedata.count('\n')
            checksum = hashlib.md5(filedata.encode("utf-8")).hexdigest()
            
            filebind[1].append((counter, name, creationtime, bytesize, linescount, checksum))
            counter = counter + 1
        filetable.append(filebind)
    return filetable

def create_xlsx_workbook(headers, widths):
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    length = len(headers)
    
    wb = Workbook()
    ws = wb.active
    
    for i in range(length):
        coord = alphabet[i] + '1'
        cell = ws[coord]
        cell.value = headers[i]
        cell.alignment = widths[i]
    
    counter = 2
    letter_end = alphabet[length - 1]
    for directorydata in filetable:
        coord1 = "A" + str(counter)
        coord2 = letter_end + str(counter)
        coord = coord1 + ':' + coord2
        
        ws.merge_cells(coord)
        cell = ws[coord1]
        cell.value = "Каталог -  " + directorydata[0]
        cell.alignment = Alignment(horizontal="center")
        counter = counter + 1
        
        for filedata in directorydata[1]:
            for i in range(len(filedata)):
                coord = alphabet[i] + str(counter)
                cell = ws[coord]
                cell.value = filedata[i]
                cell.alignment = widths[i]
            counter = counter + 1
    
    return wb

path = get_path()
print("Get all filenames...")
filenames = get_all_files_data(path)

print("Get all files...")
filetable = create_xlsx_data(path)

print("Generate xlsx file...")
headers = ["№ пп", "Имя файла", "Дата создания",
           "Размер (байт)", "Кол. строк", "Контрольная сумма"]
widths = [Alignment(horizontal="center"),
          Alignment(horizontal="general"),
          Alignment(horizontal="center"),
          Alignment(horizontal="center"),
          Alignment(horizontal="center"),
          Alignment(horizontal="general")]

wb = create_xlsx_workbook(headers, widths)
wb.save("report.xlsx")
